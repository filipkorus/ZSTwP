from openpyxl import load_workbook
import xml.etree.ElementTree as ET
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def split_text_into_lines(text, max_chars):
    """
    Split text into lines, each containing no more than max_chars characters.
    """
    words = text.split()
    lines = []
    current_line = ''
    
    for word in words:
        if len(current_line) + len(word) + 1 <= max_chars:
            current_line += ' ' + word if current_line else word
        else:
            lines.append(current_line)
            current_line = word
    
    if current_line:
        lines.append(current_line)
    
    return lines

def create_pdf(xml_file, pdf_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    c = canvas.Canvas(pdf_file, pagesize=letter)
    y = 750

    for tmf_application in root.findall('tmf-application'):
        name = tmf_application.find('name').text
        documentation = tmf_application.find('documentation').text
        
        name_lines = split_text_into_lines(f"Name: {name}", 90)
        for line in name_lines:
            c.drawString(60, y, line)
            y -= 12

        y -= 6

        documentation_lines = split_text_into_lines(documentation, 90)
        for line in documentation_lines:
            if y < 50:
                c.showPage()
                y = 750
            c.drawString(60, y, line)
            y -= 12

        y -= 6

        if y < 50:
            c.showPage()
            y = 750

    c.save()

create_pdf('requirements_data.xml', 'output.pdf')
try:
    workbook = load_workbook(filename='GB929F_v2350.xlsx')
except FileNotFoundError:
    print("File not found.")
    exit()
except Exception as e:
    print("An error occurred:", e)
    exit()
    
sheet = workbook['TAM23.5']
root = ET.Element("data")

for i in range(2,1056):
    value = sheet.cell(row=i, column=4).value
    next_value = sheet.cell(row=i + 1, column=4).value if i < 1055 else None
    
    if value != next_value:
        item = ET.SubElement(root, "tmf-application")
        title = ET.SubElement(item, "name")
        title.text = str(sheet.cell(row=i + 1, column=1).value)
        item_value = ET.SubElement(item, "documentation")
        item_value.text = str(value)
        
tree = ET.ElementTree(root)
tree.write("requirements_data.xml")

create_pdf('requirements_data.xml', 'output.pdf')